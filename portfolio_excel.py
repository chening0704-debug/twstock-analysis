# portfolio_excel.py — Excel 庫存讀取與結果回寫（v3，支援張/股單位）
import pandas as pd
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from db import upsert, get_conn, query

EXCEL_PATH  = Path("portfolio.xlsx")
SHEET_INPUT = "庫存輸入"
DATA_START  = 4   # 資料從第 4 列開始


def _convert_to_shares_float(qty, unit) -> float:
    """
    將數量 + 單位轉換為統一的「張數（浮點）」
    張 → 直接用（例：5張 = 5.0）
    股 → 除以 1000（例：300股 = 0.3張）
    """
    try:
        qty_f = float(str(qty).replace(",", "").strip())
        unit_str = str(unit).strip() if unit else "張"
        if "股" in unit_str:
            return qty_f / 1000.0   # 股數轉張數
        else:
            return qty_f            # 已經是張數
    except (TypeError, ValueError):
        return None


def load_portfolio_from_excel() -> int:
    """讀取 Excel 庫存表，驗證後寫入 my_portfolio"""
    if not EXCEL_PATH.exists():
        logging.warning(f"[Portfolio] 找不到 {EXCEL_PATH}，跳過")
        return 0

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        if SHEET_INPUT not in wb.sheetnames:
            logging.error(
                f"[Portfolio] 找不到工作表「{SHEET_INPUT}」"
            )
            return 0
        ws = wb[SHEET_INPUT]
    except Exception as e:
        logging.error(f"[Portfolio] 開啟 Excel 失敗：{e}")
        return 0

    records, errors = [], []

    for row in ws.iter_rows(
        min_row=DATA_START, values_only=True
    ):
        # A=代號 B=成本 C=數量 D=單位 E=名稱(系統) ...
        if len(row) < 4:
            continue

        stock_id   = row[0]
        cost_price = row[1]
        qty        = row[2]
        unit       = row[3]

        # 空白列跳過
        if stock_id is None or str(stock_id).strip() == "":
            continue

        # 代號清洗
        stock_id = str(stock_id).strip().zfill(4)

        # 驗證代號格式
        if not stock_id.isdigit() or not (4 <= len(stock_id) <= 5):
            errors.append(f"代號格式錯誤：{stock_id}")
            continue

        # 驗證成本
        try:
            cost_price = float(
                str(cost_price).replace(",", "").strip()
            )
            if cost_price <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(f"{stock_id} 成本錯誤：{cost_price}")
            continue

        # 驗證數量 + 單位，轉換為張數
        shares = _convert_to_shares_float(qty, unit)
        if shares is None or shares <= 0:
            errors.append(
                f"{stock_id} 數量錯誤：{qty}（單位：{unit}）"
            )
            continue

        # 從系統欄位嘗試讀取股票名稱（E欄）
        stock_name = stock_id
        if len(row) >= 5 and row[4] is not None:
            name_val = str(row[4]).strip()
            if name_val and name_val not in ["系統自動查詢","系統自動","系統填入"]:
                stock_name = name_val

        records.append({
            "stock_id":   stock_id,
            "stock_name": stock_name,
            "cost_price": round(cost_price, 2),
            "shares":     round(shares, 4),
            "buy_date":   None,
            "market":     "TSE",
            "note":       f"原始：{qty}{unit or '張'}",
        })

        logging.info(
            f"[Portfolio] 載入 {stock_id} "
            f"成本={cost_price} "
            f"數量={qty}{unit or '張'} "
            f"= {shares:.4f}張"
        )

    if errors:
        for e in errors:
            logging.warning(f"  [Portfolio] 略過：{e}")

    if not records:
        logging.warning("[Portfolio] Excel 無有效庫存資料")
        return 0

    # 清空舊資料，寫入新資料
    try:
        with get_conn() as conn:
            conn.execute("DELETE FROM my_portfolio")
    except Exception as e:
        logging.error(f"[Portfolio] 清空舊資料失敗：{e}")

    upsert("my_portfolio", pd.DataFrame(records), pk=["stock_id"])
    logging.info(
        f"[Portfolio] 成功載入 {len(records)} 支持倉"
        + (f"，略過 {len(errors)} 筆錯誤" if errors else "")
    )
    return len(records)


def write_results_to_excel(portfolio_advice: list) -> None:
    """將分析結果回寫到 Excel F~J 欄（配合新版欄位）"""
    if not EXCEL_PATH.exists() or not portfolio_advice:
        return

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_INPUT]
    except Exception as e:
        logging.error(f"[Portfolio] 回寫開啟失敗：{e}")
        return

    advice_map = {a["stock_id"]: a for a in portfolio_advice}

    signal_colors = {
        "加碼": ("E8F8F5", "196F3D"),
        "持有": ("EBF5FB", "1A5276"),
        "觀察": ("F2F3F4", "555555"),
        "減碼": ("FEF9E7", "7D6608"),
        "出清": ("FEF0E7", "784212"),
        "停損": ("FDEDEC", "922B21"),
    }
    signal_label = {
        "加碼": "🟢 加碼",
        "持有": "🔵 持有",
        "觀察": "⚪ 觀察",
        "減碼": "🟡 減碼",
        "出清": "🟠 出清",
        "停損": "🔴 停損",
    }

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    for row in ws.iter_rows(min_row=DATA_START):
        cell_a = row[0]
        if cell_a.value is None:
            continue
        sid = str(cell_a.value).strip().zfill(4)
        if sid not in advice_map:
            continue

        adv    = advice_map[sid]
        sig    = adv.get("signal", "觀察")
        bg, fg = signal_colors.get(sig, ("F2F3F4", "555555"))
        rn     = cell_a.row

        # E：股票名稱（系統查詢）
        e = ws[f"E{rn}"]
        e.value     = adv.get("name", sid)
        e.font      = Font(size=11, name="Arial", color="333333")
        e.fill      = hfill("F8F9FA")
        e.alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # F：現價
        f_c = ws[f"F{rn}"]
        f_c.value         = adv.get("current", "")
        f_c.number_format = "#,##0.00"
        f_c.font          = Font(size=11, name="Arial")
        f_c.fill          = hfill("EAFAF1")
        f_c.alignment     = Alignment(
            horizontal="center", vertical="center"
        )

        # G：損益%
        g   = ws[f"G{rn}"]
        pnl = adv.get("pnl_pct", 0) / 100
        g.value         = pnl
        g.number_format = "0.00%"
        g.font          = Font(
            bold=True, size=11, name="Arial",
            color="196F3D" if pnl >= 0 else "922B21"
        )
        g.fill      = hfill("EAFAF1" if pnl >= 0 else "FDEDEC")
        g.alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # H：綜合評分
        h       = ws[f"H{rn}"]
        h.value = round(adv.get("score", 0), 1)
        h.font  = Font(bold=True, size=11, name="Arial")
        h.fill  = hfill("EBF5FB")
        h.alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # I：RF 漲機率
        i_c       = ws[f"I{rn}"]
        rf        = adv.get("rf_prob", 0.5)
        i_c.value = f"{rf*100:.0f}%"
        i_c.font  = Font(size=11, name="Arial")
        i_c.fill  = hfill("EBF5FB")
        i_c.alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # J：操作建議
        j       = ws[f"J{rn}"]
        j.value = signal_label.get(sig, sig)
        j.font  = Font(
            bold=True, size=11, name="Arial", color=fg
        )
        j.fill      = hfill(bg)
        j.alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # 最後一列：更新時間
    last = DATA_START + 32
    ws[f"A{last}"] = "最後更新"
    ws[f"B{last}"] = pd.Timestamp.today().strftime(
        "%Y-%m-%d %H:%M"
    )
    for cell in [ws[f"A{last}"], ws[f"B{last}"]]:
        cell.font = Font(
            italic=True, size=9,
            color="888888", name="Arial"
        )

    try:
        wb.save(EXCEL_PATH)
        logging.info("[Portfolio] 結果回寫完成")
    except Exception as e:
        logging.error(f"[Portfolio] 回寫失敗：{e}")
